# -*- coding: utf-8 -*-
# @Time     : 2021-11-17 9:19
# @Author   : Chardman
# @Software : Pycharm

import datetime

# 1.新建excel
from openpyxl import Workbook

wb = Workbook()  # 实例化工作簿
ws = wb.active  # 激活worksheet

wb.save(filename='openpyxl创建的表.xlsx')

# 2.打开已存在的excel
from openpyxl import load_workbook

wb = load_workbook('openpyxl创建的表.xlsx')
ws = wb.active  # 激活worksheet
ws.title = 'my_sheet'  # 修改sheet名称

# 3.写入数据
# 方式一：数据可以直接分配到单元格中(可以输入公式)
ws['A1'] = 42
# 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
ws.append([1, 2, 3])
# 方式三：Python 类型会被自动转换
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
wb.save(filename='openpyxl创建的表.xlsx')

# 4.创建新表
wb = load_workbook('openpyxl创建的表.xlsx')
ws1 = wb.create_sheet('another_sheet')
ws2 = wb.create_sheet('third_sheet', 0)

# 5.选择表
ws3 = wb['my_sheet']
ws4 = wb.get_sheet_by_name('my_sheet')  # 过时了
print(ws3 is ws4)

# 6.查看表名
print(wb.sheetnames)
# 显示所有表名
for sheet in wb:
    print(sheet.title)
# 遍历表名

# 7.访问单元格
# 方法一
c = ws['A4']
# 方法二：row 行；column 列
d = ws.cell(row=4, column=2, value=10)
# 方法三：只要访问就创建
for i in range(1, 101):
    for j in range(1, 101):
        ws.cell(row=i, column=j)

# 通过切片
cell_range = ws['A1':'C2']
# 通过行(列)
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

# 通过指定范围(行 → 行)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# 通过指定范围(列 → 列)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# 遍历所有 方法一
ws = wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)
# 遍历所有 方法二
tuple(ws.columns)
wb.save('openpyxl创建的表.xlsx')

# 9.改变sheet标签的颜色
wb = load_workbook('openpyxl创建的表.xlsx')
ws = wb['my_sheet']
ws.sheet_properties.tabColor = '1072BA'
# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)

# 获取每一行每一列
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)

from openpyxl.utils import get_column_letter, column_index_from_string

# 根据列的数字返回字母
print(get_column_letter(2))  # B
# 根据字母返回列的数字
print(column_index_from_string('D'))  # 4

# 删除工作表
# 方式一
wb.remove(wb['third_sheet'])
# 方式二
del wb['another_sheet']
wb.save('openpyxl创建的表.xlsx')

# 10.设置单元格风格
# 需要导入的类
from openpyxl.styles import Font, colors, Alignment

wb = load_workbook('openpyxl创建的表.xlsx')
sheet = wb.active
font1 = Font(name='楷体', size=24, italic=True, color=colors.BLUE, bold=True)

sheet['A1'].font = font1
# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('B1:G1')  # 合并一行中的几个单元格
sheet.merge_cells('A1:C3')  # 合并一个矩形区域中的单元格

# 拆分单元格，拆分后，值回到A1位置
sheet.unmerge_cells('A1:C3')
wb.save('openpyxl创建的表.xlsx')

# 11. 汇总
import datetime
from random import choice
from time import time
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# 设置文件 mingc
addr = "openpyxl汇总.xlsx"
# 打开文件
try:
    wb = load_workbook(addr)
except FileNotFoundError:
    wb = Workbook()
# 创建一张新表
ws = wb.create_sheet('汇总')
# 第一行输入
ws.append(['TIME', 'TITLE', 'A-Z'])

# 输入内容（500行数据）
for i in range(500):
    TIME = datetime.datetime.now().strftime("%H:%M:%S")
    TITLE = str(time())
    A_Z = get_column_letter(choice(range(1, 50)))
    ws.append([TIME, TITLE, A_Z])

# 获取最大行
row_max = ws.max_row
# 获取最大列
con_max = ws.max_column
# 把上面写入内容打印在控制台
for j in ws.rows:  # we.rows 获取每一行数据
    for n in j:
        print(n.value, end="\t")  # n.value 获取单元格的值
    print()
# 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
wb.save(addr)

# 12. 画图
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook(write_only=True)
ws = wb.create_sheet()

rows = [
    ('Number', 'Batch 1', 'Batch 2'),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30),
]

for row in rows:
    ws.append(row)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length (mm)'

data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, "A10")

from copy import deepcopy

chart2 = deepcopy(chart1)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horizontal Bar Chart"
ws.add_chart(chart2, "G10")

chart3 = deepcopy(chart1)
chart3.type = "col"
chart3.style = 12
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = 'Stacked Chart'
ws.add_chart(chart3, "A27")

chart4 = deepcopy(chart1)
chart4.type = "bar"
chart4.style = 13
chart4.grouping = "percentStacked"
chart4.overlap = 100
chart4.title = 'Percent Stacked Chart'
ws.add_chart(chart4, "G27")

wb.save("bar.xlsx")
